#!/usr/bin/env python3
"""
BACKTEST REPORT GENERATOR

Generate comprehensive Excel reports from backtest results:
- Summary metrics
- Trade log
- Equity curve
- Parameter comparison
- Charts and visualizations
"""

import json
import sys
from datetime import datetime
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import LineChart, Reference, BarChart
except ImportError:
    print("Error: openpyxl not installed")
    print("Run: pip install openpyxl --break-system-packages")
    sys.exit(1)


class BacktestReportGenerator:
    """
    Generate Excel reports from backtest results
    """
    
    def __init__(self, results_file: str):
        """
        Args:
            results_file: Path to backtest JSON results
        """
        # Load results
        with open(results_file, 'r') as f:
            self.data = json.load(f)
        
        # Create workbook
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Define styles
        self.define_styles()
    
    def define_styles(self):
        """Define cell styles"""
        # Header style
        self.header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        # Title style
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.title_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Metric style
        self.metric_font = Font(name='Arial', size=10)
        self.positive_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        self.negative_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Border
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_summary_sheet(self, strategy: str):
        """Create summary metrics sheet"""
        ws = self.wb.create_sheet(f"{strategy.upper()} Summary")
        
        result = self.data['results'][strategy]
        metrics = result['metrics']
        
        # Title
        ws['A1'] = f"{strategy.upper()} Strategy - Backtest Summary"
        ws['A1'].font = self.title_font
        ws['A1'].fill = self.title_fill
        ws.merge_cells('A1:D1')
        
        # Period
        ws['A3'] = "Backtest Period:"
        ws['B3'] = f"{result['period']['start']} to {result['period']['end']}"
        ws['B3'].font = Font(bold=True)
        
        # Parameters
        ws['A4'] = "Parameters:"
        ws['B4'] = f"Volume Multiplier: {result['params']['volume_multiplier']}x"
        ws['B5'] = f"RSI Threshold: {result['params']['rsi_threshold']}"
        
        # Metrics header
        ws['A7'] = "Metric"
        ws['B7'] = "Value"
        ws['C7'] = "Status"
        
        for col in ['A7', 'B7', 'C7']:
            ws[col].font = self.header_font
            ws[col].fill = self.header_fill
            ws[col].border = self.thin_border
        
        # Metrics data
        metrics_list = [
            ("Total Trades", metrics['total_trades'], ""),
            ("Winning Trades", metrics['winning_trades'], ""),
            ("Losing Trades", metrics['losing_trades'], ""),
            ("Win Rate", f"{metrics['win_rate']}%", "PASS" if metrics['win_rate'] >= 55 else "FAIL"),
            ("Total Return", f"{metrics['total_return']}%", "PASS" if metrics['total_return'] > 0 else "FAIL"),
            ("Final Capital", f"{metrics['final_capital']:,.0f} VND", ""),
            ("Avg Profit", f"{metrics['avg_profit']}%", ""),
            ("Avg Loss", f"{metrics['avg_loss']}%", ""),
            ("Max Profit", f"{metrics['max_profit']}%", ""),
            ("Max Loss", f"{metrics['max_loss']}%", ""),
            ("Profit Factor", f"{metrics['profit_factor']:.2f}", "PASS" if metrics['profit_factor'] > 1.5 else "FAIL"),
            ("Expectancy", f"{metrics['expectancy']}%", ""),
            ("Avg Hold Days", f"{metrics['avg_hold_days']:.1f}", ""),
            ("Max Drawdown", f"{metrics['max_drawdown']}%", "PASS" if metrics['max_drawdown'] > -15 else "FAIL")
        ]
        
        row = 8
        for metric, value, status in metrics_list:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'C{row}'] = status
            
            # Apply borders
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = self.thin_border
            
            # Color code status
            if status == "PASS":
                ws[f'C{row}'].fill = self.positive_fill
                ws[f'C{row}'].font = Font(bold=True, color='006100')
            elif status == "FAIL":
                ws[f'C{row}'].fill = self.negative_fill
                ws[f'C{row}'].font = Font(bold=True, color='9C0006')
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
    
    def create_trades_sheet(self, strategy: str):
        """Create detailed trades log"""
        ws = self.wb.create_sheet(f"{strategy.upper()} Trades")
        
        result = self.data['results'][strategy]
        trades = result['trades']
        
        if not trades:
            ws['A1'] = "No trades executed"
            return
        
        # Convert to DataFrame
        df = pd.DataFrame(trades)
        
        # Headers
        headers = ['Code', 'Entry Date', 'Entry Price', 'Exit Date', 'Exit Price', 
                   'Shares', 'Profit', 'Profit %', 'Hold Days', 'Exit Reason']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        
        # Data rows
        for row_idx, trade in enumerate(trades, 2):
            ws.cell(row=row_idx, column=1, value=trade['code'])
            ws.cell(row=row_idx, column=2, value=str(trade['entry_date']))
            ws.cell(row=row_idx, column=3, value=f"{trade['entry_price']:,.0f}")
            ws.cell(row=row_idx, column=4, value=str(trade['exit_date']))
            ws.cell(row=row_idx, column=5, value=f"{trade['exit_price']:,.0f}")
            ws.cell(row=row_idx, column=6, value=trade['shares'])
            ws.cell(row=row_idx, column=7, value=f"{trade['profit']:,.0f}")
            ws.cell(row=row_idx, column=8, value=f"{trade['profit_pct']:.2f}%")
            ws.cell(row=row_idx, column=9, value=trade['hold_days'])
            ws.cell(row=row_idx, column=10, value=trade['exit_reason'])
            
            # Color code profit
            profit_cell = ws.cell(row=row_idx, column=8)
            if trade['profit'] > 0:
                profit_cell.fill = self.positive_fill
                profit_cell.font = Font(color='006100')
            else:
                profit_cell.fill = self.negative_fill
                profit_cell.font = Font(color='9C0006')
            
            # Apply borders
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).border = self.thin_border
        
        # Adjust widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 15
    
    def create_equity_curve_sheet(self, strategy: str):
        """Create equity curve visualization"""
        ws = self.wb.create_sheet(f"{strategy.upper()} Equity")
        
        result = self.data['results'][strategy]
        equity_data = result['equity_curve']
        
        if not equity_data:
            ws['A1'] = "No equity data"
            return
        
        # Headers
        ws['A1'] = 'Date'
        ws['B1'] = 'Equity'
        ws['C1'] = 'Open Positions'
        
        for col in ['A1', 'B1', 'C1']:
            ws[col].font = self.header_font
            ws[col].fill = self.header_fill
            ws[col].border = self.thin_border
        
        # Data
        for row_idx, point in enumerate(equity_data, 2):
            ws.cell(row=row_idx, column=1, value=str(point['date']))
            ws.cell(row=row_idx, column=2, value=point['equity'])
            ws.cell(row=row_idx, column=3, value=point['open_positions'])
        
        # Create chart
        chart = LineChart()
        chart.title = f"{strategy.upper()} Equity Curve"
        chart.y_axis.title = 'Equity (VND)'
        chart.x_axis.title = 'Date'
        
        # Data reference
        data = Reference(ws, min_col=2, min_row=1, max_row=len(equity_data)+1)
        chart.add_data(data, titles_from_data=True)
        
        # Position chart
        ws.add_chart(chart, "E2")
        
        # Adjust widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def create_comparison_sheet(self):
        """Create strategy comparison"""
        ws = self.wb.create_sheet("Comparison", 0)  # First sheet
        
        # Title
        ws['A1'] = "Strategy Comparison - Backtest Results"
        ws['A1'].font = self.title_font
        ws['A1'].fill = self.title_fill
        ws.merge_cells('A1:D1')
        
        # Headers
        ws['A3'] = "Metric"
        ws['B3'] = "Breakout"
        ws['C3'] = "Divergence"
        ws['D3'] = "Winner"
        
        for col in ['A3', 'B3', 'C3', 'D3']:
            ws[col].font = self.header_font
            ws[col].fill = self.header_fill
            ws[col].border = self.thin_border
        
        # Get metrics
        b_metrics = self.data['results']['breakout']['metrics']
        d_metrics = self.data['results']['divergence']['metrics']
        
        # Comparison data
        comparisons = [
            ("Win Rate", f"{b_metrics['win_rate']}%", f"{d_metrics['win_rate']}%"),
            ("Total Return", f"{b_metrics['total_return']}%", f"{d_metrics['total_return']}%"),
            ("Profit Factor", f"{b_metrics['profit_factor']:.2f}", f"{d_metrics['profit_factor']:.2f}"),
            ("Max Drawdown", f"{b_metrics['max_drawdown']}%", f"{d_metrics['max_drawdown']}%"),
            ("Total Trades", str(b_metrics['total_trades']), str(d_metrics['total_trades'])),
            ("Avg Profit", f"{b_metrics['avg_profit']}%", f"{d_metrics['avg_profit']}%"),
            ("Avg Loss", f"{b_metrics['avg_loss']}%", f"{d_metrics['avg_loss']}%")
        ]
        
        row = 4
        for metric, breakout_val, divergence_val in comparisons:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = breakout_val
            ws[f'C{row}'] = divergence_val
            
            # Determine winner (for positive metrics)
            if metric in ["Win Rate", "Total Return", "Profit Factor", "Avg Profit"]:
                breakout_num = float(breakout_val.replace('%', '').replace(',', ''))
                divergence_num = float(divergence_val.replace('%', '').replace(',', ''))
                
                if breakout_num > divergence_num:
                    ws[f'D{row}'] = "Breakout"
                    ws[f'D{row}'].fill = self.positive_fill
                else:
                    ws[f'D{row}'] = "Divergence"
                    ws[f'D{row}'].fill = self.positive_fill
            
            # Apply borders
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = self.thin_border
            
            row += 1
        
        # Adjust widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def generate(self, output_file: str):
        """
        Generate complete report
        
        Args:
            output_file: Output Excel filename
        """
        print(f"Generating report: {output_file}", file=sys.stderr)
        
        # Create comparison sheet
        self.create_comparison_sheet()
        
        # Create sheets for each strategy
        for strategy in ['breakout', 'divergence']:
            self.create_summary_sheet(strategy)
            self.create_trades_sheet(strategy)
            self.create_equity_curve_sheet(strategy)
        
        # Save workbook
        self.wb.save(output_file)
        
        print(f"✅ Report saved: {output_file}", file=sys.stderr)


def main():
    """Main execution"""
    if len(sys.argv) < 2:
        print("Usage: python generate_report.py <backtest_results.json>")
        sys.exit(1)
    
    results_file = sys.argv[1]
    output_file = results_file.replace('.json', '_report.xlsx')
    
    # Generate report
    generator = BacktestReportGenerator(results_file)
    generator.generate(output_file)
    
    print(f"\nReport generated: {output_file}")


if __name__ == '__main__':
    main()
